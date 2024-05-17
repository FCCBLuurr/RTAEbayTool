import os
import json
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from icecream import ic  # For better debug outputs

def update_spreadsheet(file_path):
    base_dir = '/Users/ejrta/Documents/Coding Folders/Scripts and Macros/Tools and Workflow Scripts/Flickr&Ebay/components/(extract)'
    # Load JSON data
    temp_dir = os.path.join(base_dir, 'temp')
    temp_data_path = os.path.join(temp_dir, 'temp.json')

    with open(temp_data_path, 'r') as file:
        temp_data = json.load(file)

    ic("Keys in temp_data:", temp_data.keys())  # Debug: Print keys in temp_data dictionary

    # Load the workbook
    workbook = load_workbook(filename=file_path)
    sheet = workbook['Inventory Sheet']

    # Find the table range manually or define it explicitly
    table_range = 'A4:X1000'  # Update this range to match the actual table range in your worksheet

    # Extract the start and end row indices from the table range
    table_rows = table_range.split(':')
    if len(table_rows) == 2:  # Table range spans only one row
        table_start_row = table_end_row = int(table_rows[0][1:])
    else:  # Table range spans multiple rows
        table_start_row, _, table_end_row = table_rows
        table_start_row = int(table_start_row[1:])
        table_end_row = int(table_end_row[1:])

    ic("Table range:", table_range)

    # Find the last row with data in the SKU and Listed? columns within the table range
    last_row_B = max((idx for idx, cell in enumerate(sheet['B'][table_start_row-1:table_end_row], start=table_start_row) if cell.value is not None), default=table_start_row-2)
    last_row_W = max((idx for idx, cell in enumerate(sheet['W'][table_start_row-1:table_end_row], start=table_start_row) if cell.value is not None), default=table_start_row-2)

    ic("Last row with data in column SKU:", last_row_B)
    ic("Last row with data in column Listed?:", last_row_W)

    # Iterate over keys in temp_data and update entry in column Listed? to "Y" if SKU matches and entry is "N"
    for sku in temp_data.keys():
        if sku in [str(cell.value) for cell in sheet['B'][table_start_row-1:last_row_B+1]] and temp_data[sku]["Stats"]["Listed?"] == "N":
            ic("SKU found in spreadsheet:", sku)
            for cell in sheet['B']:
                if cell.value == int(sku):
                    ic("Updating row:", cell.row)
                    sheet.cell(row=cell.row, column=23, value='Y')  # Update Listed? column

    # Check if all rows with "N" exist in temp.json
    all_listed = all(temp_data.get(sku, {}).get("Stats", {}).get("Listed?", "") == "Y" for sku in temp_data.keys())

    if all_listed:
        workbook.save(file_path)  # Save the workbook if all rows with "N" are now "Y"
        ic("All rows with 'N' are now 'Y'. Spreadsheet saved.")
    else:
        ic("Not all rows with 'N' exist in temp.json. Spreadsheet not saved.")

if __name__ == '__main__':
    root = tk.Tk()
    root.withdraw()
    initial_dir = '/users/ejrta/Documents/Excel Sheets/Ebay Inventory Sheet'
    file_path = filedialog.askopenfilename(
        title="Select the Excel file",
        initialdir=initial_dir,
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    root.destroy()

    if file_path:
        update_spreadsheet(file_path)
    else:
        print("No file selected.")
